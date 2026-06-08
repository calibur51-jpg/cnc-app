import streamlit as st
import pandas as pd
import requests
import time
import io
from PIL import Image
import streamlit as st
import datetime

# 1. 這裡處理電腦瀏覽器分頁的小圖示 (使用本地路徑)
st.set_page_config(
    page_title="明星刀管", 
    page_icon="icon.png",  
    layout="wide"
)

# 2. 這裡處理手機「加入主畫面」的專屬圖示 (保留 URL)
# 因為 HTML 注入相對路徑容易跑掉，保留 URL 反而比較穩定
ICON_URL = "https://raw.githubusercontent.com/calibur51-jpg/cnc-app/main/icon.png?v=4"

st.markdown(f"""
    <meta name="mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
    <meta name="apple-mobile-web-app-title" content="明星刀管">
    <link rel="apple-touch-icon" href="{ICON_URL}">
    <link rel="shortcut icon" href="{ICON_URL}">
""", unsafe_allow_html=True)


# --- 【初始化】自動化 QR 參數讀取 ---
if 'scanned_id' not in st.session_state:
    st.session_state.scanned_id = None

# 讀取網址參數 (自動偵測掃描結果)
query_params = st.query_params
if "scan" in query_params:
    st.session_state.scanned_id = query_params["scan"]

# --- 1. 設定區 ---
INV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTo2vi_36qF4mzPkxzNOJPTip7y-TXJLBm745noRRa4v_L_qkJ0DhFkaJ7tvYLCYWdFV3wbXOtH--zJ/pub?gid=0&single=true&output=csv"
LOG_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTo2vi_36qF4mzPkxzNOJPTip7y-TXJLBm745noRRa4v_L_qkJ0DhFkaJ7tvYLCYWdFV3wbXOtH--zJ/pub?gid=1320901506&single=true&output=csv"
SET_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTo2vi_36qF4mzPkxzNOJPTip7y-TXJLBm745noRRa4v_L_qkJ0DhFkaJ7tvYLCYWdFV3wbXOtH--zJ/pub?gid=657176737&single=true&output=csv"

WEBHOOK_URL = "https://script.google.com/macros/s/AKfycbyZ6-S7x4fp4iCQbdpClMlXQFUxQ9q036XFtCZxuObS2mqaF7wv-U26QOJhqGsvxHyskQ/exec"

# --- 2. 函數區 ---
# 💡 修正 1：加上 ttl=5（快取5秒自動過期），並指定 show_spinner=False 讓重整更流暢
@st.cache_data(ttl=5, show_spinner=False)
def get_data():
    try:
        ts = time.time()
        df_inv = pd.read_csv(f"{INV_URL}&_={ts}", encoding='utf-8-sig')
        df_log = pd.read_csv(f"{LOG_URL}&_={ts}", encoding='utf-8-sig')
        df_set = pd.read_csv(f"{SET_URL}&_={ts}", encoding='utf-8-sig')
        return df_inv, df_log, df_set
    except Exception as e:
        st.error(f"連線失敗: {e}")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

def post_data_to_sheet(payload):
    try:
        response = requests.post(WEBHOOK_URL, json=payload)
        return response.status_code == 200
    except:
        return False

# --- 3. 畫面開頭與天王老子同步按鈕 ---
st.title("🛠️ CNC 刀具庫存管理系統")

# 💡 修正 2：在最頂端直接提供手動重擊鈕，強制洗掉快取與 session 記憶
if st.button("🔄 立即同步雲端試算表 (當現場資料卡住時點擊)"):
    st.cache_data.clear()              # 清除所有快取
    get_data.clear()                   # 專門強制清除 get_data 函式的快取
    if "data" in st.session_state:
        del st.session_state.data      # 徹底刪除記憶體舊骨架
    st.success("✅ 已成功抹除舊記憶！正在重新連線 Google Sheets...")
    st.rerun()

st.divider()

# --- 4. 資料安全載入流 ---
# 💡 修正 3：確保載入邏輯能正確連動重新載入
if "data" not in st.session_state:
    st.session_state.data = get_data()

df_inv, df_log, df_set = st.session_state.data

t1, t2, t3, t4 = st.tabs(["領用", "後台", "紀錄", "進貨與盤點系統"])
with t1:
    st.header("🔪 刀具領用")
    _, df_log, df_set = st.session_state.data
    
    # [這裡我把最上方的檢查刪掉了，避免訊息跑回最上面]

    # 1. 處理掃描邏輯 (維持原樣)
    default_cat_idx = 0
    if "scanned_id" in st.session_state and st.session_state.scanned_id is not None:
        match = df_inv[df_inv["刀具編號"].astype(str) == st.session_state.scanned_id]
        if not match.empty:
            cat_name = match.iloc[0]["分類"]
            all_cats = ["全部"] + df_inv["分類"].unique().tolist()
            if cat_name in all_cats:
                default_cat_idx = all_cats.index(cat_name)
            st.session_state.pending_tool = match.iloc[0]["品名規格"]
        else:
            st.error(f"❌ 找不到編號 {st.session_state.scanned_id} 的刀具")
        st.session_state.scanned_id = None 
    
    all_cats = ["全部"] + df_inv["分類"].unique().tolist()
    cat_sel = st.selectbox("分類", all_cats, index=default_cat_idx, key="t1_cat")
    
    df_f = df_inv if cat_sel == "全部" else df_inv[df_inv["分類"] == cat_sel]
    t_list = df_f["品名規格"].tolist()
    
    tool_idx = 0
    if "pending_tool" in st.session_state:
        if st.session_state.pending_tool in t_list:
            tool_idx = t_list.index(st.session_state.pending_tool)
        del st.session_state.pending_tool
        
    t_name = st.selectbox("選擇領用刀具", t_list, index=tool_idx, key="t1_tool")

    cur_stock = 0
    tool_info = df_inv[df_inv["品名規格"] == t_name]
    if not tool_info.empty:
        idx = tool_info.index[0]
        t_sel = tool_info.loc[idx, "刀具編號"]
        cur_stock = int(tool_info.loc[idx, "架上"])
   try:
        warehouse_stock = int(float(tool_info["倉庫數量"]))
   except:
        warehouse_stock = 0
        st.info(f"編號:{t_sel} | 架上:{cur_stock} 支")
    else:
        st.warning("⚠️ 請選擇刀具規格")

    if "q_val" not in st.session_state: st.session_state["q_val"] = 1
    c1, c2 = st.columns(2)
    with c1: 
        if st.button("➕ 加1", key="b_add"): 
            st.session_state["q_val"] += 1
            st.rerun()
    with c2: 
        if st.button("🔄 歸1", key="b_res"): 
            st.session_state["q_val"] = 1
            st.rerun()
            
    qty = st.number_input("數量", min_value=1, value=st.session_state["q_val"])
    st.session_state["q_val"] = qty
    u_list = df_set["人員"].replace("", pd.NA).dropna().unique().tolist()
    u = st.selectbox("人員", u_list, key="t1_user")
    m = "無"
    r = st.selectbox("原因", ["正常消耗 (磨損/架機)", "異常損耗 (斷刀/其他)"], key="t1_reason")
    wo = st.text_input("工單", key="t1_wo").strip()
    
    msg_area = st.empty()
    
    if st.button("確認領用", type="primary", use_container_width=True):
        if qty > cur_stock:
            msg_area.error("❌ 庫存不足！")
        else:
            with st.spinner("正在執行領用作業..."):
                payload = {"action": "領用", "row": idx + 2, "t_sel": t_sel, "qty": qty, "u": u, "m": m, "r": r, "wo": wo}
                try:
                    response = requests.post(WEBHOOK_URL, json=payload, timeout=10)
                    if response.status_code == 200:
                        st.session_state.data[0].loc[idx, "目前庫存"] -= qty
                        st.session_state["q_val"] = 1
                        # 成功訊息儲存至記憶體
                        st.session_state["notify_msg"] = f"✅ 已領刀：{t_name} x {qty}"
                        st.rerun() 
                    else:
                        msg_area.error(f"❌ 寫入失敗")
                except Exception as e:
                    msg_area.error(f"❌ 寫入失敗: {e}")

    # --- 【這裡才是正確的位置】 ---
    # 只有當程式跑到這裡時，才會檢查有沒有訊息，所以訊息會顯示在按鈕下方
    if "notify_msg" in st.session_state:
        msg_area.success(st.session_state["notify_msg"]) # 使用 msg_area 會更精準定位在按鈕區下方
        del st.session_state["notify_msg"]

with t2:
    st.header("🔒 管理員專區")
    pw = st.text_input("輸入管理員密碼", type="password", key="pw_t2")
    
    if pw == "1234":
        st.success("✅ 驗證成功，進入管理模式")
        st.divider()
        with st.expander("📦 點擊查看所有刀具庫存總覽", expanded=False):
            categories = ["全部"] + df_inv["分類"].dropna().unique().tolist()
            c1, c2 = st.columns(2)
            with c1:
                sel_cat = st.selectbox("分類篩選", options=categories, key="t2_view_cat")
            with c2:
                search_text = st.text_input("搜尋關鍵字 (編號/名稱)", key="t2_view_search")
            df_show = df_inv.copy()
            if sel_cat != "全部":
                df_show = df_show[df_show["分類"] == sel_cat]
            if search_text:
                mask = df_show["刀具編號"].astype(str).str.contains(search_text, case=False, na=False) | \
                       df_show["品名規格"].astype(str).str.contains(search_text, case=False, na=False)
                df_show = df_show[mask]
            st.dataframe(df_show, use_container_width=True)
        st.divider()
        st.header("⚙️ 系統管理")
        mode = st.radio("選擇操作模式", ["刀具建檔", "庫存校正"], horizontal=True)
        
        all_categories = ["全部"] + df_inv["分類"].dropna().unique().tolist()
        
        # --- 1. 刀具建檔區塊 ---
        if mode == "刀具建檔":
            st.subheader("📝 新增刀具")
            if st.session_state.get("last_action") == "建檔":
                st.success("✅ 建檔成功！系統已同步。")
                del st.session_state.last_action
            
            category_options = df_inv["分類"].dropna().unique().tolist()
            if "其他" not in category_options:
                category_options.append("其他")
                
            with st.form("new_tool_form", clear_on_submit=True):
                new_id = st.text_input("刀具編號")
                new_name = st.text_input("品名規格")
                new_cat = st.selectbox("分類", options=category_options)
                new_qty = st.number_input("初始庫存", min_value=0, value=0)
                if st.form_submit_button("確認建檔"):
                    payload = {"action": "建檔", "t_id": new_id, "t_name": new_name, "cat": new_cat, "loc": new_loc, "qty": new_qty}
                    if post_data_to_sheet(payload):
                        st.session_state.last_action = "建檔"
                        st.rerun()
                    else:
                        st.error("❌ 建檔失敗")
                        
        # --- 2. 庫存校正區塊 ---
        elif mode == "庫存校正":
            st.subheader("🔧 庫存數量校正")
            if st.session_state.get("last_action") == "校正":
                st.success("✅ 庫存校正成功！")
                del st.session_state.last_action
                
            cc1, cc2 = st.columns(2)
            with cc1:
                adj_cat = st.selectbox("篩選分類", options=all_categories, key="t2_adj_cat")
            with cc2:
                adj_search = st.text_input("關鍵字搜尋", key="t2_adj_search")
                
            df_adj_show = df_inv.copy()
            if adj_cat != "全部":
                df_adj_show = df_adj_show[df_adj_show["分類"] == adj_cat]
            if adj_search:
                mask_adj = df_adj_show["刀具編號"].astype(str).str.contains(adj_search, case=False, na=False) | \
                           df_adj_show["品名規格"].astype(str).str.contains(adj_search, case=False, na=False)
                df_adj_show = df_adj_show[mask_adj]
                
            tool_list_t2 = df_adj_show["品名規格"].tolist()
            
            if not tool_list_t2:
                st.error("❌ 找不到符合條件的刀具，請重新篩選或清除關鍵字")
            else:
                target_tool = st.selectbox("選擇目標刀具", tool_list_t2, key="t2_tool_sel")
                matched_inv_df = df_inv[df_inv["品名規格"] == target_tool]
                
                if not matched_inv_df.empty:
                    current_inv = matched_inv_df.iloc[0]
                    
                    try:
                        raw_qty = int(pd.Series(current_inv['currently_stock']).to_numeric(errors='coerce').fillna(0).astype(int).iloc[0])
                    except:
                        try:
                            raw_qty = int(pd.Series(current_inv['目前庫存']).to_numeric(errors='coerce').fillna(0).astype(int).iloc[0])
                        except:
                            raw_qty = 0
                        
                    default_qty = max(0, raw_qty) 
                    st.info(f"📊 目前系統紀錄庫存：{raw_qty}支 ]")
                    
                    new_adj_qty = st.number_input("輸入正確現場庫存總數", min_value=0, value=default_qty)
                    
                    if st.button("確認校正", key="t2_adj_confirm_btn"):
                        payload = {"action": "校正", "t_sel": current_inv['刀具編號'], "new_qty": new_adj_qty}
                        if post_data_to_sheet(payload):
                            st.session_state.last_action = "校正"
                            idx = df_inv[df_inv["刀具編號"] == current_inv['刀具編號']].index[0]
                            try:
                                st.session_state.data[0].loc[idx, "目前庫存"] = new_adj_qty
                            except:
                                st.session_state.data[0].loc[idx, "currently_stock"] = new_adj_qty
                                
                            st.cache_data.clear() 
                            st.rerun()
                        else:
                            st.error("❌ 校正失敗")
                else:
                    st.warning("⚠️ 刀具資料同步中，請稍候...")

    elif pw != "":
        st.warning("⚠️ 密碼錯誤，請重新輸入")
    else:
        st.info("請輸入管理員密碼以存取管理功能")
with t3:
    st.header("📊 刀具管理戰情室")
    pw = st.text_input("輸入管理員密碼", type="password", key="pw_t3")
    
    if pw == "1234":
        st.cache_data.clear()
        _, df_log, _ = st.session_state.data
        df_inv = st.session_state.data[0]
        
        # 1. 精裝 Excel 格式化函式
        def get_styled_excel(df, title):
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=title, index=False)
                ws = writer.sheets[title]
                header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                header_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
                for col_num in range(1, len(df.columns) + 1):
                    ws.cell(row=1, column=col_num).fill = header_fill
                    ws.cell(row=1, column=col_num).font = header_font
            return buffer

        if not df_log.empty:
            # --- 【關鍵修復】：強效處理日期格式 ---
            df_log["時間"] = df_log["時間"].astype(str).str.replace("下午", " PM").str.replace("上午", " AM")
            df_log["時間"] = pd.to_datetime(df_log["時間"], errors='coerce')
            df_log["月份"] = df_log["時間"].dt.strftime("%Y-%m").fillna("未分類")
            df_log["數量"] = pd.to_numeric(df_log["數量"], errors='coerce').fillna(0)
            df_log["價格"] = pd.to_numeric(df_log["價格"], errors='coerce').fillna(0)
            
            current_month = datetime.datetime.now().strftime("%Y-%m")

            # --- 1. 本月對帳收折 ---
            with st.expander("📅 本月財務與進銷存對帳", expanded=True):
                df_this_month = df_log[df_log["月份"] == current_month].copy()
                if not df_this_month.empty:
                    df_in = df_this_month[df_this_month["動作"] == "進貨"].groupby("刀具編號")["數量"].sum().reset_index(name="本月進貨量")
                    df_out = df_this_month[df_this_month["動作"] == "領用"].groupby("刀具編號")["數量"].sum().reset_index(name="本月領用量")
                    
                    # --- [修正點]：加入"架上"欄位並確保讀取正確 ---
                    df_acc = df_inv[["分類", "刀具編號", "品名規格", "架上", "倉庫數量"]].merge(df_in, on="刀具編號", how="left").merge(df_out, on="刀具編號", how="left").fillna(0)
                    
                    # 復刻：備用單價映射與財務計算
                    inv_price_map = pd.to_numeric(df_inv.set_index("刀具編號")["單價"], errors='coerce').fillna(0).to_dict()
                    df_acc["當月單價"] = df_acc["刀具編號"].map(inv_price_map).fillna(0)
                    df_acc["本月新購買總金額"] = df_acc["本月進貨量"] * df_acc["當月單價"]
                    
                    # --- [修正點]：將架上與倉庫加總後計算價值 ---
                    df_acc["現有庫存總價值"] = (df_acc["架上"] + df_acc["倉庫數量"]) * df_acc["當月單價"]
                    
                    st.metric("本月新購總額", f"${int(df_acc['本月新購買總金額'].sum()):,}")
                    st.dataframe(df_acc, use_container_width=True)
                    st.download_button("📥 下載本月精裝報表", get_styled_excel(df_acc, "本月對帳").getvalue(), f"對帳表_{current_month}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                else:
                    st.info(f"本月 ({current_month}) 目前尚無資料")

            # --- 2. 歷史總計分析 ---
            with st.expander("📊 歷史總計與篩選"):
                view_mode = st.radio("資料範圍選擇:", ["全時段總計", "僅本月"], horizontal=True)
                df_view = df_log if view_mode == "全時段總計" else df_log[df_log["月份"] == current_month].copy()
                df_usage = df_view[df_view["動作"] == "領用"].copy()
                
                c1, c2 = st.columns(2)
                with c1: st.markdown("**人員領用排行**"); st.bar_chart(df_usage.groupby("經辦人員")["數量"].sum())
                with c2: st.markdown("**原因分析**"); st.bar_chart(df_usage.groupby("原因類型")["數量"].sum())
                
                st.subheader("📜 歷史紀錄進階篩選")
                search_wo = st.text_input("🔍 搜尋工單號碼:")
                if search_wo: df_view = df_view[df_view["工單號碼"].astype(str).str.contains(search_wo, case=False, na=False)]
                
                st.dataframe(df_view.sort_values(by="時間", ascending=False), use_container_width=True)
                st.download_button("📥 下載歷史紀錄精裝版", get_styled_excel(df_view, "紀錄").getvalue(), "歷史紀錄.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    elif pw != "":
        st.warning("⚠️ 密碼錯誤")
with t4:
    st.header("📥 進貨與盤點系統")
    pw = st.text_input("輸入管理員密碼", type="password", key="pw_t4")
    
    if pw == "1234":
        st.success("✅ 驗證成功")
        st.divider()
        
        # --- [獨立功能：架上庫存快查] ---
        st.subheader("📦 架上補貨檢查")
        df_inv["架上"] = pd.to_numeric(df_inv["架上"], errors='coerce').fillna(0)
        low_shelf_df = df_inv[df_inv["架上"] < 2].copy()
        
        if not low_shelf_df.empty:
            st.warning(f"🚨 注意：共有 {len(low_shelf_df)} 項刀具「架上數量低於 2」，請盡快補貨！")
            st.dataframe(low_shelf_df[["品名規格", "架上"]], use_container_width=True, hide_index=True)
        else:
            st.success("✅ 架上庫存充足，無需補貨")
        # -------------------------------

        st.divider()
        st.subheader("⚙️ 選擇目標刀具")
        
        categories = ["全部"] + df_inv["分類"].dropna().unique().tolist()
        c1, c2 = st.columns(2)
        with c1: sel_cat = st.selectbox("篩選分類", options=categories, key="t4_cat")
        with c2: search_text = st.text_input("關鍵字搜尋", key="t4_search")
            
        df_show = df_inv.copy()
        if sel_cat != "全部": df_show = df_show[df_show["分類"] == sel_cat]
        if search_text:
            mask = df_show["刀具編號"].astype(str).str.contains(search_text, case=False, na=False) | \
                   df_show["品名規格"].astype(str).str.contains(search_text, case=False, na=False)
            df_show = df_show[mask]
            
        tool_options = df_show["品名規格"].tolist()
        if not tool_options:
            st.error("找不到符合條件的刀具")
        else:
            sel_tool_name = st.selectbox("搜尋結果", tool_options, key="t4_tool_sel")
            matched_df = df_show[df_show["品名規格"] == sel_tool_name]
            
            if not matched_df.empty:
                tool_info = matched_df.iloc[0]
                
                # --- [修補]：強制轉換，處理空值導致的 int() 報錯 ---
                cur_shelf = int(pd.to_numeric(tool_info["架上"], errors='coerce') or 0)
                cur_wh = int(pd.to_numeric(tool_info["倉庫數量"], errors='coerce') or 0)
                # ---------------------------------------------
                
                try: current_price = float(tool_info["單價"])
                except: current_price = 0.0
                    
                col_a, col_b, col_c = st.columns(3)
                col_a.metric("架上", cur_shelf)
                col_b.metric("倉庫", cur_wh)
                col_c.metric("單價", f"${int(current_price)}")
                
                mode = st.radio("選擇操作模式", ["進貨", "上架", "盤點"], horizontal=True)
                
                with st.form("t4_action_form", clear_on_submit=True):
                    qty_input = st.number_input("數量", min_value=0, value=0)
                    price_input = st.number_input("本次單價", min_value=0.0, value=current_price, step=10.0) if mode == "進貨" else current_price
                    u_input = st.text_input("操作人員")
                    
                    if st.form_submit_button(f"確認{mode}"):
                        payload = {
                            "action": mode,
                            "t_id": tool_info["刀具編號"],
                            "qty": qty_input,
                            "price": price_input,
                            "u": u_input
                        }
                        
                        if post_data_to_sheet(payload):
                            idx = df_inv[df_inv["刀具編號"] == tool_info["刀具編號"]].index[0]
                            
                            # 強制轉型修復 TypeError
                            df = st.session_state.data[0]
                            df["架上"] = pd.to_numeric(df["架上"], errors='coerce').fillna(0)
                            df["倉庫數量"] = pd.to_numeric(df["倉庫數量"], errors='coerce').fillna(0)
                            
                            # 處理記憶體更新
                            if mode == "進貨":
                                df.loc[idx, "倉庫數量"] = cur_wh + qty_input
                                df.loc[idx, "單價"] = price_input
                            elif mode == "上架":
                                df.loc[idx, "架上"] = cur_shelf + qty_input
                                df.loc[idx, "倉庫數量"] = max(0, cur_wh - qty_input)
                            elif mode == "盤點":
                                df.loc[idx, "倉庫數量"] = qty_input
                            
                            st.cache_data.clear() 
                            st.session_state.success_msg = f"✅ {mode}成功！"
                            st.rerun()
                        else:
                            st.error("❌ 操作失敗")
            else:
                st.warning("⚠️ 刀具資料加載中...")
                
        if "success_msg" in st.session_state:
            st.success(st.session_state.success_msg)
            del st.session_state.success_msg
                
        if "success_msg" in st.session_state:
            st.success(st.session_state.success_msg)
            del st.session_state.success_msg
